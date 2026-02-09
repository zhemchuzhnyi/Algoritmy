import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# Данные
dates = [
    '12.05.2025', '13.05.2025', '14.05.2025', '15.05.2025', '16.05.2025', '17.05.2025', '18.05.2025',
    '19.05.2025', '20.05.2025', '21.05.2025', '22.05.2025', '23.05.2025', '24.05.2025', '25.05.2025'
]
requests = [
    110, 121, 113, 134, 114, 72, 67,  # Умеренный режим (сумма = 731)
    85, 78, 76, 76, 86, 56, 68       # Редкий режим (сумма = 525)
]
modes = ['Умеренный']*7 + ['Редкий']*7

# Создаем DataFrame
df = pd.DataFrame({
    'Дата': dates,
    'Заявки': requests,
    'Режим': modes
})

# Создаем Excel-файл
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Заявки Докма"

# Заголовки
headers = ['Дата', 'Заявки', 'Режим']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color='4B9CFF', end_color='4B9CFF', fill_type='solid')

# Заполняем данные
for row, (date, request, mode) in enumerate(zip(dates, requests, modes), 2):
    ws.cell(row=row, column=1).value = date
    ws.cell(row=row, column=2).value = request
    ws.cell(row=row, column=3).value = mode
    # Форматирование: цвет фона в зависимости от режима
    fill_color = '4B9CFF' if mode == 'Умеренный' else '34C759'
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                     top=Side(style='thin'), bottom=Side(style='thin'))

# Автонастройка ширины столбцов
for col in range(1, 4):
    column_letter = get_column_letter(col)
    ws.column_dimensions[column_letter].width = 15

# Добавляем диаграмму
chart = BarChart()
chart.title = "Количество заявок на сервис Докма (12.05–25.05.2025)"
chart.x_axis.title = "Дата"
chart.y_axis.title = "Количество заявок"
chart.legend = None  # Без легенды, так как цвета в таблице уже показывают режимы

# Данные для диаграммы
data = Reference(ws, min_col=2, min_row=2, max_row=15, max_col=2)
cats = Reference(ws, min_col=1, min_row=2, max_row=15)
chart.add_data(data)
chart.set_categories(cats)

# Настраиваем цвета столбцов (синий для умеренного, зелёный для редкого)
for i, series in enumerate(chart.series):
    series.graphicalProperties.solidFill = '4B9CFF' if i < 7 else '34C759'

# Позиция диаграммы
ws.add_chart(chart, "E2")


# Сохраняем файл
wb.save('applications_data.xlsx')